import logging
import os
from typing import Optional, List

from django.core.files.uploadedfile import UploadedFile
from langchain.chains.summarize import load_summarize_chain
from langchain.prompts import (
    ChatPromptTemplate,
    HumanMessagePromptTemplate,
    SystemMessagePromptTemplate,
)
import fitz  # PyMuPDF
from openpyxl import load_workbook
from langchain_core.documents import Document
from langchain_core.language_models import BaseChatModel
from langchain_core.output_parsers import PydanticOutputParser
from langchain_openai import ChatOpenAI
from ..schemas import TaskPlan, TaskPlanListContainer # Added TaskPlanListContainer

logger = logging.getLogger(__name__)


class ServiceManualsAssistantService:
    """Service for processing service manual documents and extracting structured data."""

    def __init__(self):
        self.temp_file_path: Optional[str] = None
        self.temp_dir: str = "temp_documents"
        os.makedirs(self.temp_dir, exist_ok=True)

    def process_document(
        self, file: UploadedFile, llm: BaseChatModel
    ) -> List[TaskPlan]: 
        """
        Process a service manual document using LangChain and generate structured output.

        Args:
            file: The uploaded PDF or XLSX file to process
            llm: The language model to use for processing

        Returns:
            List[TaskPlan]: A list of structured task plans extracted from document

        Raises:
            ValueError: If file format is not supported
            IOError: If file cannot be read or processed
        """
        file_extension = file.name.lower().split('.')[-1]
        
        if file_extension not in ["pdf", "xlsx"]:
            raise ValueError("Only PDF and XLSX files are supported")

        try:
            file_path = self._save_temp_file(file)
            self.temp_file_path = file_path

            # Extract text based on file type
            if file_extension == "pdf":
                full_text = self._extract_text_from_pdf(file_path)
            elif file_extension == "xlsx":
                full_text = self._extract_text_from_xlsx(file_path)

            # Create a single LangChain Document
            documents = [Document(page_content=full_text)]

            if not documents or not documents[0].page_content.strip():
                logger.warning(f"No content extracted from {file.name}")
                return None

            # Pass file extension to processing method to use appropriate prompt
            processed_data_container = self._process_task_plan_data(llm, documents, file_extension)
            return processed_data_container.task_plans # Extract list from container

        except Exception as e:
            logger.error(f"Error processing document {file.name}: {str(e)}")
            raise
        finally:
            self.cleanup()

    def _save_temp_file(self, file: UploadedFile) -> str:
        """
        Save the uploaded file to a temporary location.
        Args:
            file: The uploaded file to save
        Returns:
            str: Path to the saved temporary file
        """
        file_path = os.path.join(self.temp_dir, file.name)
        with open(file_path, "wb+") as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        logger.info(f"Saved temporary file: {file_path}")
        return file_path

    def _extract_text_from_pdf(self, file_path: str) -> str:
        """
        Extract text from a PDF file.
        
        Args:
            file_path: Path to the PDF file
            
        Returns:
            str: Extracted text content
        """
        doc = fitz.open(file_path)
        full_text = ""
        for page in doc:
            full_text += page.get_text()
        doc.close()
        return full_text

    def _extract_text_from_xlsx(self, file_path: str) -> str:
        """
        Extract text from an XLSX file and convert to markdown format.
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            str: Extracted text content in markdown format
        """
        workbook = load_workbook(file_path, data_only=True)
        markdown_content = []
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Add sheet header
            markdown_content.append(f"# {sheet_name}\n")
            
            # Find the actual data range
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            if max_row == 1 and max_col == 1:
                # Empty sheet
                markdown_content.append("*No data in this sheet*\n")
                continue
            
            # Process the data
            current_table = []
            current_task_plan = None
            
            for row in worksheet.iter_rows(min_row=1, max_row=max_row, 
                                         min_col=1, max_col=max_col, values_only=True):
                # Skip empty rows
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    if current_table:
                        # Process current table before moving to next section
                        markdown_content.extend(self._process_table_to_markdown(current_table))
                        current_table = []
                    continue
                
                # Convert None values to empty strings for processing
                row_data = [str(cell) if cell is not None else "" for cell in row]
                current_table.append(row_data)
            
            # Process any remaining table data
            if current_table:
                markdown_content.extend(self._process_table_to_markdown(current_table))
            
            markdown_content.append("\n---\n")  # Sheet separator
        
        workbook.close()
        return "\n".join(markdown_content)

    def _process_table_to_markdown(self, table_data: List[List[str]]) -> List[str]:
        """
        Convert table data to markdown format, handling task plans and checklists.
        
        Args:
            table_data: List of rows, each row is a list of cell values
            
        Returns:
            List[str]: Markdown formatted lines
        """
        if not table_data:
            return []
        
        markdown_lines = []
        
        # Check if this looks like a structured task plan table
        header_row = table_data[0]
        
        # Common patterns for task plan tables
        task_plan_indicators = ["task", "plan", "code", "description", "checklist", "step", "procedure"]
        is_task_table = any(any(indicator in str(cell).lower() for indicator in task_plan_indicators) 
                           for cell in header_row if cell)
        
        if is_task_table and len(table_data) > 1:
            # Process as a structured task plan table
            markdown_lines.append("## Task Plan Details\n")
            
            # Create markdown table
            headers = [str(cell).strip() for cell in header_row]
            non_empty_headers = [h for h in headers if h]
            
            if non_empty_headers:
                # Add table headers
                header_line = "| " + " | ".join(non_empty_headers) + " |"
                separator_line = "|" + "|".join([" --- " for _ in non_empty_headers]) + "|"
                
                markdown_lines.extend([header_line, separator_line])
                
                # Add data rows
                for row in table_data[1:]:
                    # Take only as many cells as we have headers
                    row_cells = [str(cell).strip() if cell else "" for cell in row[:len(non_empty_headers)]]
                    row_line = "| " + " | ".join(row_cells) + " |"
                    markdown_lines.append(row_line)
                
                markdown_lines.append("")  # Empty line after table
        else:
            # Process as general content
            for i, row in enumerate(table_data):
                non_empty_cells = [str(cell).strip() for cell in row if cell and str(cell).strip()]
                
                if non_empty_cells:
                    if i == 0 and len(non_empty_cells) == 1:
                        # Single cell in first row might be a heading
                        markdown_lines.append(f"## {non_empty_cells[0]}\n")
                    elif len(non_empty_cells) == 1:
                        # Single cell content
                        markdown_lines.append(f"- {non_empty_cells[0]}")
                    else:
                        # Multiple cells - format as a list or table row
                        if len(non_empty_cells) == 2:
                            markdown_lines.append(f"**{non_empty_cells[0]}:** {non_empty_cells[1]}")
                        else:
                            markdown_lines.append("- " + " | ".join(non_empty_cells))
        
        return markdown_lines

    def _create_task_plan_prompt(self, file_extension: str) -> ChatPromptTemplate: 
        """
        Create the prompt template for task plan data extraction.
        Args:
            file_extension: The file extension to determine which prompt to use
        Returns:
            ChatPromptTemplate: The configured prompt template
        """
        system_template = self._get_system_prompt_template(file_extension)
        system_message_prompt = SystemMessagePromptTemplate.from_template(system_template)
        
        if file_extension == "xlsx":
            human_template = """Please analyze the following Excel/spreadsheet document and extract the required information:
            
            {text}
            
            IMPORTANT: This is data from an Excel spreadsheet. Use the EXACT codes and IDs that appear in the spreadsheet columns.
            Do NOT generate new codes or IDs - use only the ones already present in the table data.
            
            Respond with ONLY the JSON structure containing the extracted task plan data, matching the schema provided in the system instructions.
            """
        else:
            human_template = """Please analyze the following service manual document and extract the required information:
            
            {text}
            
            Respond with ONLY the JSON structure containing the extracted task plan data, matching the schema provided in the system instructions.
            """
        
        human_message_prompt = HumanMessagePromptTemplate.from_template(human_template)
        return ChatPromptTemplate.from_messages([system_message_prompt, human_message_prompt])

    def _get_system_prompt_template(self, file_extension: str) -> str:
        """
        Get the system prompt template for task plan data extraction.
        Args:
            file_extension: The file extension to determine which prompt to use
        Returns:
            str: The system prompt template
        """
        base_instructions = """You are an expert AI system designed to transform equipment service manuals, technical bulletins, and historical repair logs into standardized work orders.
Your primary goal is to generate structured Task Plans and detailed Checklists for equipment troubleshooting and maintenance.

Key Objectives:
1.  **Ingest Technical Documentation:** Analyze the provided text from equipment manuals or related documents.
2.  **Identify Maintenance Procedures:** Recognize patterns indicating troubleshooting sequences, repair steps, safety precautions, and required resources.
3.  **Extract Contextual Relationships:** Understand links between symptoms, potential causes, and corresponding resolution actions.
4.  **Prioritize Safety:** Identify and clearly call out all safety warnings, cautions, required Personal Protective Equipment (PPE), and safety-critical steps.

Output Requirements:
-   Generate standardized Task Plans in a logical, step-by-step format.
-   Create comprehensive Checklists associated with each Task Plan. Checklists MUST include:
    -   Required tools and parts.
    -   All relevant safety procedures and PPE requirements.
    -   Clear, sequential troubleshooting or maintenance instructions.
    -   Verification steps to confirm successful completion.
-   Organize tasks considering safety, logical flow, and operational efficiency."""

        if file_extension == "xlsx":
            code_instructions = """
CRITICAL INSTRUCTION FOR EXCEL/SPREADSHEET DATA:
-   **USE EXACT CODES FROM SPREADSHEET:** When processing Excel/spreadsheet data, you MUST use the exact codes and IDs that appear in the spreadsheet columns.
-   **DO NOT GENERATE NEW CODES:** Never create new task codes or checklist IDs. Only use the ones already present in the table data.
-   **PRESERVE ORIGINAL VALUES:** If you see codes like 'TP_123', 'CL_ABC', etc. in the spreadsheet, use those exact values.
-   **MATCH TABLE STRUCTURE:** Follow the structure and relationships shown in the spreadsheet tables.

For spreadsheet data, codes and IDs should be:
-   **Exact matches:** Use the precise codes found in the spreadsheet columns (e.g., if you see 'TP_123' in a task code column, use 'TP_123').
-   **Preserved formatting:** Maintain any formatting, capitalization, or special characters from the original codes.
-   **Consistent with source:** Ensure the extracted codes exactly match what appears in the source spreadsheet."""
        else:
            code_instructions = """
For each entity (Task Plan, Checklist), generate codes and IDs that are:
-   Descriptive: Incorporate key terms or abbreviations from the equipment, procedure, or task (e.g., 'PUMP_REPLACE_SEAL', 'CHK_BEARING_INSPECT').
-   Unique: Ensure each code/ID is distinct, potentially using a combination of the manual's name, task type, and sequence.
-   Human-readable: Prefer meaningful identifiers over random strings."""

        return """{{base_instructions}}

{{code_instructions}}

Format your response STRICTLY as a raw JSON object with a single key "task_plans".
The value of "task_plans" should be a list, where each item in the list is a Task Plan object.
Do NOT use any markdown formatting (e.g., do NOT use ```json ... ```).
The output must be directly parsable as JSON. 

The schema for the overall JSON object is:
{{
    "task_plans": [
        {{
            "task_code": string, // Task plan code, descriptive and unique (e.g., 'TP_PUMP_OVERHAUL') max length 20
            "description": string, // Task plan description (e.g., 'Complete overhaul of Pump Model X') max length 80
            "checklist": [
                {{
                    "checklist_id": string, // Checklist ID, descriptive and unique (e.g., 'CL_DISASSEMBLY') max length 20
                    "description": string // Checklist item description (e.g., 'Step 1: Verify pump is isolated and LOTO is applied') max length 80
                }},
                {{
                    "checklist_id": string, // (e.g., 'CL_INSPECT_SHAFT') max length 20
                    "description": string // (e.g., 'Step 2: Inspect shaft for wear and runout. Required tools: Micrometer') max length 80
                }}
                // ... more checklist items ...
            ]
        }}
        // ... more task plans if applicable ...
    ]
}}

Extract all relevant task plans and their checklist steps from the document.
If specific details (like tool names, part numbers) are mentioned for a step, include them in the checklist item's description.
Ensure every step is actionable and clear for a maintenance technician.
For any missing required fields in the schema (task_code, description, checklist_id), use reasonable, descriptive defaults based on the content.
""".replace("{{base_instructions}}", base_instructions).replace("{{code_instructions}}", code_instructions)

    def _process_task_plan_data(
        self, llm: ChatOpenAI, documents: list[Document], file_extension: str
    ) -> TaskPlanListContainer: # Changed return type to the container
        """
        Process the document chunks to extract task plan information.
        Args:
            llm: The language model to use for processing
            documents: The document chunks to process
            file_extension: The file extension to determine which prompt to use
        Returns:
            TaskPlanListContainer: Container with structured task plan data
        """
        logger.info(f"Processing {len(documents)} document chunks for {file_extension} file")
        chat_prompt = self._create_task_plan_prompt(file_extension)
        parser = PydanticOutputParser(pydantic_object=TaskPlanListContainer) # Use the container model

        summarize_chain = load_summarize_chain(
            llm,
            chain_type="stuff",
            verbose=True,
            prompt=chat_prompt,
        )

        try:
            response = summarize_chain.invoke({"input_documents": documents})
            output_text = response["output_text"]
            # Strip markdown JSON fences if present - keeping this as a safeguard
            if output_text.startswith("```json\n"):
                output_text = output_text[7:]
            if output_text.endswith("\n```"):
                output_text = output_text[:-4]
            
            parsed_response = parser.invoke(output_text)
            logger.info("Successfully extracted task plan data container")
            return parsed_response
        except Exception as e:
            logger.error(f"Error extracting task plan data: {str(e)}")
            raise ValueError(f"Failed to extract structured data from document: {str(e)}")

    def cleanup(self) -> None:
        """Clean up temporary files."""
        if self.temp_file_path and os.path.exists(self.temp_file_path):
            try:
                os.remove(self.temp_file_path)
                logger.info(f"Cleaned up temporary file: {self.temp_file_path}")
            except Exception as e:
                logger.error(f"Error cleaning up temporary file: {str(e)}") 